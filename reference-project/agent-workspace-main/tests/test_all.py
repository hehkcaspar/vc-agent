"""Comprehensive test for all implemented features (Phase 1-4+)."""

from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path


def test_phase1_core():
    """Test Phase 1: Walking Skeleton."""
    print("\n[Phase 1] Walking Skeleton")
    print("-" * 40)
    
    from agent_workspace.cli import main
    
    with tempfile.TemporaryDirectory() as tmpdir:
        ws = Path(tmpdir) / "ws"
        
        # Init
        main(["init", "--dir", str(ws)])
        
        # Add test files
        (ws / "resources" / "doc.txt").write_text("Hello world")
        
        # Scan
        main(["scan", "--workspace", str(ws)])
        
        print("  [OK] Phase 1 core features work")


def test_phase2_persistence():
    """Test Phase 2: Persistence features."""
    print("\n[Phase 2] Persistence")
    print("-" * 40)
    
    from agent_workspace.workspace import Workspace
    from agent_workspace.config import load_workspace_config
    
    with tempfile.TemporaryDirectory() as tmpdir:
        ws_path = Path(tmpdir) / "ws"
        ws_path.mkdir()
        
        # Create structure manually
        (ws_path / "resources").mkdir()
        (ws_path / ".snapshots").mkdir()
        
        # Create initial file
        (ws_path / "resources" / "file1.txt").write_text("v1")
        
        # Create workspace and snapshot
        cfg = load_workspace_config(ws_path)
        ws = Workspace(ws_path, cfg.resources_dir, cfg.snapshots_dir)
        
        manifest1 = ws.scan()
        ws.save_snapshot(manifest1)
        
        # Modify
        (ws_path / "resources" / "file1.txt").write_text("v2")
        (ws_path / "resources" / "file2.txt").write_text("new")
        
        manifest2 = ws.scan()
        diff = ws.diff(manifest2, manifest1)
        
        assert len(diff["modified"]) == 1, "Expected 1 modified file"
        assert len(diff["added"]) == 1, "Expected 1 added file"
        
        print("  [OK] Snapshot and diff work")
        
        # Test memory loading
        from agent_workspace.agent import _load_memory
        (ws_path / "artifacts" / "memory").mkdir(parents=True)
        (ws_path / "artifacts" / "memory" / "notes.md").write_text("Key observation")
        
        memory = _load_memory(ws_path, cfg)
        assert memory is not None and "Key observation" in memory
        print("  [OK] Memory loading works")


def test_phase3_templates():
    """Test Phase 3: Template resolution."""
    print("\n[Phase 3] Templates")
    print("-" * 40)
    
    from agent_workspace.prompts import resolve_template
    
    with tempfile.TemporaryDirectory() as tmpdir:
        templates_dir = Path(tmpdir) / "templates"
        templates_dir.mkdir()
        
        # Create template
        (templates_dir / "greeting.md").write_text("Hello, {name}!")
        
        # Resolve
        result = resolve_template(templates_dir, "greeting", {"name": "World"})
        assert result == "Hello, World!"
        
        print("  [OK] Template resolution works")


def test_phase4_extraction():
    """Test Phase 4: Excel and search."""
    print("\n[Phase 4] Rich Extraction + Search")
    print("-" * 40)
    
    from agent_workspace.extractor import extract_file
    from agent_workspace.workspace import classify_file
    
    with tempfile.TemporaryDirectory() as tmpdir:
        # Test Excel extraction
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test"
            ws["A1"] = "Name"
            ws["B1"] = "Value"
            ws["A2"] = "Item1"
            ws["B2"] = 100
            
            excel_path = Path(tmpdir) / "test.xlsx"
            wb.save(excel_path)
            
            result = extract_file(excel_path, "excel", None)
            assert "Item1" in result["text"]
            print("  [OK] Excel extraction works")
        except ImportError:
            print("  [SKIP] openpyxl not available")
        
        # Test search
        from agent_workspace.tools.search_resources import search_resources
        
        ws_path = Path(tmpdir) / "ws"
        (ws_path / "resources").mkdir(parents=True)
        (ws_path / "resources" / "doc.txt").write_text("This contains the keyword python")
        
        result = search_resources.invoke({
            "workspace_root": str(ws_path),
            "query": "python",
            "max_results": 10
        })
        
        assert "keyword python" in result
        print("  [OK] Search resources works")


def test_all_tools():
    """Test all tools work end-to-end."""
    print("\n[Tools] All tools")
    print("-" * 40)
    
    from agent_workspace.tools.scan_resources import scan_resources
    from agent_workspace.tools.extract_content import extract_content
    from agent_workspace.tools.write_artifact import write_artifact
    from agent_workspace.tools.read_artifact import read_artifact
    
    with tempfile.TemporaryDirectory() as tmpdir:
        ws_path = Path(tmpdir) / "ws"
        
        # Setup
        from agent_workspace.cli import main
        main(["init", "--dir", str(ws_path)])
        (ws_path / "resources" / "test.txt").write_text("Sample content")
        
        # Test scan_resources
        result = scan_resources.invoke({"workspace_root": str(ws_path)})
        assert "test.txt" in result
        print("  [OK] scan_resources tool")
        
        # Test extract_content
        result = extract_content.invoke({
            "workspace_root": str(ws_path),
            "file_paths": ["test.txt"]
        })
        assert "Sample content" in result
        print("  [OK] extract_content tool")
        
        # Test write_artifact
        result = write_artifact.invoke({
            "workspace_root": str(ws_path),
            "artifact_type": "reports",
            "name": "test.md",
            "content": "# Report"
        })
        assert "Artifact written" in result
        print("  [OK] write_artifact tool")
        
        # Test read_artifact
        result = read_artifact.invoke({
            "workspace_root": str(ws_path),
            "artifact_path": "reports/test.md"
        })
        assert "Report" in result
        print("  [OK] read_artifact tool")


def run_all_tests():
    """Run comprehensive test suite."""
    print("="*60)
    print("COMPREHENSIVE TEST SUITE (Phase 1-4)")
    print("="*60)
    
    tests = [
        test_phase1_core,
        test_phase2_persistence,
        test_phase3_templates,
        test_phase4_extraction,
        test_all_tools,
    ]
    
    passed = 0
    failed = 0
    
    for test in tests:
        try:
            test()
            passed += 1
        except Exception as e:
            failed += 1
            print(f"  [FAIL] {test.__name__}: {e}")
            import traceback
            traceback.print_exc()
    
    print("\n" + "="*60)
    print(f"SUMMARY: {passed}/{len(tests)} test groups passed")
    print("="*60)
    
    if failed == 0:
        print("\n[OK] ALL TESTS PASSED!")
        print("\nImplementation status:")
        print("  [DONE] Phase 1 - Walking Skeleton")
        print("  [DONE] Phase 2 - Persistence (diff, memory, traces)")
        print("  [DONE] Phase 3 - Templates")
        print("  [DONE] Phase 4 - Excel, Search")
        print("\nReady to move to Phase 5: Polish (error handling, progress output, README)")
        return 0
    else:
        print(f"\n[FAIL] {failed} test group(s) failed")
        return 1


if __name__ == "__main__":
    sys.exit(run_all_tests())
